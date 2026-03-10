"""XLSX cell styles and formatting.

Defines reusable styles for both sheets:
- Header rows: dark background, white bold text
- Data rows: alternating light/white for readability
- Pivot apartment totals: accent background, bold
- Pivot grand total: dark background, white bold
- Column widths auto-fitted to content
- Freeze panes on header rows
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# -- Colors ------------------------------------------------------------------

_DARK_BG = "2F5496"  # Dark blue
_ACCENT_BG = "D6E4F0"  # Light blue
_ALT_ROW_BG = "F2F2F2"  # Very light grey
_WHITE = "FFFFFF"
_BLACK = "000000"

# -- Borders -----------------------------------------------------------------

_THIN_SIDE = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

# -- Fonts -------------------------------------------------------------------

_FONT_NAME = "Aptos Narrow"

FONT_HEADER = Font(name=_FONT_NAME, bold=True, color=_WHITE, size=11)
FONT_SUBHEADER = Font(name=_FONT_NAME, bold=True, color=_BLACK, size=11)
FONT_DATA = Font(name=_FONT_NAME, color=_BLACK, size=11)
FONT_PIVOT_GROUP = Font(name=_FONT_NAME, bold=True, color=_BLACK, size=11)
FONT_PIVOT_TOTAL = Font(name=_FONT_NAME, bold=True, color=_WHITE, size=11)

# -- Fills -------------------------------------------------------------------

FILL_HEADER = PatternFill(start_color=_DARK_BG, end_color=_DARK_BG, fill_type="solid")
FILL_ACCENT = PatternFill(
    start_color=_ACCENT_BG, end_color=_ACCENT_BG, fill_type="solid"
)
FILL_ALT_ROW = PatternFill(
    start_color=_ALT_ROW_BG, end_color=_ALT_ROW_BG, fill_type="solid"
)
FILL_WHITE = PatternFill(start_color=_WHITE, end_color=_WHITE, fill_type="solid")
FILL_DARK = PatternFill(start_color=_DARK_BG, end_color=_DARK_BG, fill_type="solid")

# -- Alignment ---------------------------------------------------------------

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


_HEADER_ROW_HEIGHT = 22  # points — slightly taller than default (~15)


def set_row_height(ws, row: int, height: float = _HEADER_ROW_HEIGHT) -> None:
    """Set the height of a row in points."""
    ws.row_dimensions[row].height = height


def apply_header_style(ws, row: int, max_col: int) -> None:
    """Apply dark header style to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = _THIN_BORDER
        cell.alignment = ALIGN_CENTER
    set_row_height(ws, row)


def apply_subheader_style(ws, row: int, max_col: int) -> None:
    """Apply accent subheader style to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_SUBHEADER
        cell.fill = FILL_ACCENT
        cell.border = _THIN_BORDER
        cell.alignment = ALIGN_CENTER
    set_row_height(ws, row)


def apply_data_row_style(ws, row: int, max_col: int, alternate: bool) -> None:
    """Apply data row style with optional alternating background."""
    fill = FILL_ALT_ROW if alternate else FILL_WHITE
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_DATA
        cell.fill = fill
        cell.border = _THIN_BORDER
        cell.alignment = ALIGN_LEFT


def apply_pivot_group_style(ws, row: int, max_col: int) -> None:
    """Apply bold accent style for pivot apartment total rows."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_PIVOT_GROUP
        cell.fill = FILL_ACCENT
        cell.border = _THIN_BORDER
        cell.alignment = ALIGN_LEFT
    set_row_height(ws, row)


def apply_pivot_total_style(ws, row: int, max_col: int) -> None:
    """Apply dark bold style for pivot grand total row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_PIVOT_TOTAL
        cell.fill = FILL_DARK
        cell.border = _THIN_BORDER
        cell.alignment = ALIGN_LEFT
    set_row_height(ws, row)


def auto_fit_columns(ws, *, min_width: float = 10, max_width: float = 40) -> None:
    """Auto-fit column widths based on content.

    Scans all cells in the sheet and sets each column width to fit the
    longest value, clamped between min_width and max_width.
    """
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len

        # Add padding, clamp to range
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def freeze_at_row(ws, row: int) -> None:
    """Freeze panes so everything above the given row stays visible."""
    ws.freeze_panes = ws.cell(row=row, column=1)
