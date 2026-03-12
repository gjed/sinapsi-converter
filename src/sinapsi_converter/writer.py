"""XLSX output writer.

Produces a workbook with two sheets:
1. Raw data sheet: file header + concentrators + sorted HCA device rows
2. PIVOT sheet: apartment-grouped summary with HCA totals
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import SortCondition, SortState

from .models import ConcentratorDevice, HCADevice, ParsedReport, PivotGroup
from .styles import (
    apply_data_row_style,
    apply_header_style,
    apply_pivot_group_style,
    apply_pivot_total_style,
    apply_subheader_style,
    auto_fit_columns,
    freeze_at_row,
)


# Column headers for the file metadata section
_FILE_HEADERS = [
    "Nome file",
    "Data Report",
    "Ora Report",
    "Riferimento impianto",
    "Descrizione",
    "Totale dispositivi cablati",
    "Totale dispositivi wireless",
    "Totale dispositivi non ricevuti",
    "Totale concentratori",
]

# Column headers for concentrator devices (12 columns)
_CONCENTRATOR_HEADERS = [
    "count",
    "primary_address",
    "device_serial_number",
    "name_device",
    "device_description",
    "device_detail",
    "device_measure_hex",
    "0=wired|1=wireless|2=smart gateway",
    "model_id",
    "readout_date",
    "readout_time",
    "communication_status",
]

# Column headers for HCA devices (19 columns)
_HCA_HEADERS = [
    "count",
    "primary_address",
    "device_serial_number",
    "name_device",
    "device_description",
    "device_detail",
    "device_measure_hex",
    "0=wired M1M2|1=wireless",
    "model_id",
    "readout_date",
    "readout_time",
    "communication_status",
    "HCA",
    "HCA",
    "DATE",
    "HCA",
    "DATE",
    "DATE",
    "DATE_TIME",
]


def write_xlsx(
    report: ParsedReport,
    sorted_hca: list[HCADevice],
    pivot_groups: list[PivotGroup],
    output_path: Path,
) -> Path:
    """Write the complete XLSX workbook.

    Args:
        report: Parsed report with header and concentrators.
        sorted_hca: HCA devices sorted for the raw sheet.
        pivot_groups: Pivot groups for the summary sheet.
        output_path: Where to write the .xlsx file.

    Returns:
        The output path (for convenience).
    """
    wb = Workbook()

    # Sheet 1: PIVOT (created first so it's the default sheet, then renamed)
    ws_pivot = wb.active
    assert ws_pivot is not None
    ws_pivot.title = "PIVOT "
    _write_pivot_sheet(ws_pivot, pivot_groups)

    # Sheet 2: Raw data
    raw_sheet_name = _build_raw_sheet_name(report.header.filename)
    ws_raw = wb.create_sheet(title=raw_sheet_name)
    _write_raw_sheet(ws_raw, report, sorted_hca)

    wb.save(output_path)
    return output_path


def _build_raw_sheet_name(filename: str) -> str:
    """Build the raw sheet name from the report filename.

    Excel sheet names are limited to 31 characters.
    Prefix with 'MERC  ' to match the expected format.
    """
    name = f"MERC  {filename}"
    return name[:31]


def _write_raw_sheet(
    ws,
    report: ParsedReport,
    sorted_hca: list[HCADevice],
) -> None:
    """Write the raw data sheet."""
    max_col = len(_HCA_HEADERS)  # 19 — widest section

    # Row 1: File header labels
    for col, header in enumerate(_FILE_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    apply_header_style(ws, row=1, max_col=len(_FILE_HEADERS))

    # Row 2: File header values
    h = report.header
    row2_values = [
        h.filename,
        h.report_date,
        h.report_time,
        h.building_reference,
        h.description,
        h.total_wired,
        h.total_wireless,
        h.total_not_received,
        h.total_concentrators,
    ]
    for col, val in enumerate(row2_values, 1):
        ws.cell(row=2, column=col, value=val)
    apply_data_row_style(ws, row=2, max_col=len(_FILE_HEADERS), alternate=False)

    # Row 3: blank (separator between file header and concentrators)
    current_row = 4

    # Write concentrator devices (each with its own header row)
    for conc in report.concentrators:
        # Header row
        for col, header in enumerate(_CONCENTRATOR_HEADERS, 1):
            ws.cell(row=current_row, column=col, value=header)
        apply_subheader_style(ws, row=current_row, max_col=len(_CONCENTRATOR_HEADERS))
        current_row += 1

        # Data row
        _write_concentrator_row(ws, current_row, conc)
        apply_data_row_style(
            ws, row=current_row, max_col=len(_CONCENTRATOR_HEADERS), alternate=False
        )
        current_row += 1

        # Blank row after each concentrator
        current_row += 1

    # Blank row separating concentrators from HCA section
    current_row += 1

    # Sub-headers for HCA section
    subheader_row = current_row
    ws.cell(row=subheader_row, column=13, value="Attuale")
    ws.cell(row=subheader_row, column=14, value="Es.Prec")
    apply_subheader_style(ws, row=subheader_row, max_col=max_col)
    current_row += 1

    # HCA column headers
    hca_header_row = current_row
    for col, header in enumerate(_HCA_HEADERS, 1):
        ws.cell(row=current_row, column=col, value=header)
    apply_header_style(ws, row=current_row, max_col=max_col)
    current_row += 1

    # HCA device rows
    first_hca_row = current_row
    for idx, device in enumerate(sorted_hca):
        _write_hca_row(ws, current_row, device)
        apply_data_row_style(
            ws, row=current_row, max_col=max_col, alternate=(idx % 2 == 1)
        )
        current_row += 1

    # SUM formulas row (totals for HCA columns)
    last_hca_row = current_row - 1
    for col in (13, 14, 16):  # HCA columns M, N, P
        col_letter = chr(ord("A") + col - 1)
        formula = f"=SUM({col_letter}{first_hca_row}:{col_letter}{last_hca_row})"
        ws.cell(row=current_row, column=col, value=formula)
    apply_subheader_style(ws, row=current_row, max_col=max_col)

    # Auto-filter on the HCA table (header row through last data row)
    last_col_letter = get_column_letter(max_col)
    ws.auto_filter.ref = f"A{hca_header_row}:{last_col_letter}{last_hca_row}"

    # Sort state: primary = device_detail (col F) descending, secondary = device_description (col E) ascending
    ws.auto_filter.sortState = SortState(
        ref=f"A{first_hca_row}:{last_col_letter}{last_hca_row}",
        sortCondition=[
            SortCondition(ref=f"F{first_hca_row}:F{last_hca_row}", descending=True),
            SortCondition(ref=f"E{first_hca_row}:E{last_hca_row}"),
        ],
    )

    # Auto-fit and freeze panes below the HCA header row
    auto_fit_columns(ws)
    freeze_at_row(ws, row=hca_header_row + 1)


def _write_concentrator_row(ws, row: int, device: ConcentratorDevice) -> None:
    """Write a single concentrator device row."""
    values = [
        device.count,
        device.primary_address or None,
        device.serial_number,
        device.name,
        device.description or None,
        device.detail or None,
        device.measure_hex or None,
        device.wired_wireless,
        device.model_id,
        device.readout_date,
        device.readout_time,
        device.communication_status,
    ]
    for col, val in enumerate(values, 1):
        if val is not None and val != "":
            ws.cell(row=row, column=col, value=val)


def _write_hca_row(ws, row: int, device: HCADevice) -> None:
    """Write a single HCA device row."""
    values = [
        device.count,
        device.primary_address or None,
        _try_numeric(device.serial_number),
        device.name,
        device.description,
        device.detail,
        device.measure_hex,
        device.wired_wireless or None,
        _try_numeric(device.model_id),
        device.readout_date,
        device.readout_time,
        device.communication_status,
        device.hca_current,
        device.hca_previous_season,
        device.date_1,
        device.hca_3,
        device.date_2,
        device.date_3,
        device.datetime_1,
    ]
    for col, val in enumerate(values, 1):
        if val is not None and val != "":
            ws.cell(row=row, column=col, value=val)


def _try_numeric(value: str) -> int | str:
    """Convert string to int if it's purely numeric, otherwise return as-is."""
    try:
        return int(value)
    except (ValueError, TypeError):
        return value


def _write_pivot_sheet(ws, pivot_groups: list[PivotGroup]) -> None:
    """Write the PIVOT summary sheet."""
    max_col = 2

    # Row 3: Headers (rows 1-2 are intentionally blank, matching the sample)
    ws.cell(row=3, column=1, value="Etichette di riga")
    ws.cell(row=3, column=2, value="Somma di HCA")
    apply_header_style(ws, row=3, max_col=max_col)

    current_row = 4
    grand_total = 0.0
    device_idx = 0  # running counter for alternating rows across all groups
    current_detail: str | None = None

    for group in pivot_groups:
        # Detail section header row — emitted whenever the detail value changes
        if group.detail != current_detail:
            current_detail = group.detail
            ws.cell(row=current_row, column=1, value=group.detail)
            apply_header_style(ws, row=current_row, max_col=max_col)
            current_row += 1

        # Apartment subtotal row
        ws.cell(row=current_row, column=1, value=group.apartment)
        ws.cell(row=current_row, column=2, value=int(group.total))
        apply_pivot_group_style(ws, row=current_row, max_col=max_col)
        current_row += 1

        # Individual device rows
        for name, hca in group.devices:
            ws.cell(row=current_row, column=1, value=name)
            ws.cell(row=current_row, column=2, value=int(hca))
            apply_data_row_style(
                ws, row=current_row, max_col=max_col, alternate=(device_idx % 2 == 1)
            )
            current_row += 1
            device_idx += 1

        grand_total += group.total

    # Grand total row
    ws.cell(row=current_row, column=1, value="Totale complessivo")
    ws.cell(row=current_row, column=2, value=int(grand_total))
    apply_pivot_total_style(ws, row=current_row, max_col=max_col)

    # Auto-fit and freeze panes below the header row
    auto_fit_columns(ws)
    freeze_at_row(ws, row=4)
