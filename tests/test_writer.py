"""Tests for the XLSX writer."""

import tempfile
from pathlib import Path

import openpyxl

from sinapsi_converter.models import (
    ConcentratorDevice,
    HCADevice,
    ParsedReport,
    PivotGroup,
    ReportHeader,
)
from sinapsi_converter.writer import write_xlsx


def _make_report() -> tuple[ParsedReport, list[HCADevice], list[PivotGroup]]:
    """Create minimal test data."""
    header = ReportHeader(
        filename="RAW_report_9999_2026-01-01_09-00",
        report_date="2026-01-01",
        report_time="09:00",
        building_reference="TEST BUILDING",
        description="via test 1",
        total_wired=0,
        total_wireless=2,
        total_not_received=0,
        total_concentrators=1,
    )

    concentrator = ConcentratorDevice(
        count=0,
        primary_address="",
        serial_number="RM0001",
        name="RM0001",
        description="",
        detail="",
        measure_hex="",
        wired_wireless="2-RM0001",
        model_id="SIN.EQRPT868XM",
        readout_date="2026-01-01",
        readout_time="09:00:00",
        communication_status="Ok",
    )

    hca1 = HCADevice(
        count=1,
        primary_address="",
        serial_number="12345",
        name="ROOM A",
        description="1 TENANT INT 1",
        detail="Via Test 1",
        measure_hex="08|UDR",
        wired_wireless="",
        model_id="6251",
        readout_date="2026-01-01",
        readout_time="10:00:00",
        communication_status="Ok",
        hca_current=100.0,
        hca_previous_season=50.0,
        date_1="31/10/2025",
        hca_3=80.0,
        date_2="31/12/2025",
        date_3="31/MM/yy",
        datetime_1="01/01/2026 09:00",
    )

    hca2 = HCADevice(
        count=2,
        primary_address="",
        serial_number="12346",
        name="ROOM B",
        description="1 TENANT INT 1",
        detail="Via Test 1",
        measure_hex="08|UDR",
        wired_wireless="",
        model_id="6251",
        readout_date="2026-01-01",
        readout_time="10:05:00",
        communication_status="Ok",
        hca_current=200.0,
        hca_previous_season=150.0,
        date_1="31/10/2025",
        hca_3=180.0,
        date_2="31/12/2025",
        date_3="31/MM/yy",
        datetime_1="01/01/2026 09:05",
    )

    report = ParsedReport(
        header=header,
        concentrators=[concentrator],
        hca_devices=[hca1, hca2],
    )

    sorted_hca = [hca1, hca2]

    pivot_groups = [
        PivotGroup(
            apartment="1 TENANT INT 1",
            devices=[("ROOM A", 100.0), ("ROOM B", 200.0)],
            total=300.0,
        )
    ]

    return report, sorted_hca, pivot_groups


def test_creates_xlsx_file():
    """Writer should create an XLSX file."""
    report, sorted_hca, pivot_groups = _make_report()

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "test.xlsx"
        write_xlsx(report, sorted_hca, pivot_groups, out)
        assert out.exists()
        assert out.stat().st_size > 0


def test_has_two_sheets():
    """Output should have PIVOT and raw data sheets."""
    report, sorted_hca, pivot_groups = _make_report()

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "test.xlsx"
        write_xlsx(report, sorted_hca, pivot_groups, out)
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) == 2
        assert wb.sheetnames[0] == "PIVOT "


def test_pivot_content():
    """PIVOT sheet should contain correct groups and totals."""
    report, sorted_hca, pivot_groups = _make_report()

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "test.xlsx"
        write_xlsx(report, sorted_hca, pivot_groups, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["PIVOT "]

        # Row 3: headers
        assert ws.cell(row=3, column=1).value == "Etichette di riga"
        assert ws.cell(row=3, column=2).value == "Somma di HCA"

        # Row 4: apartment total
        assert ws.cell(row=4, column=1).value == "1 TENANT INT 1"
        assert ws.cell(row=4, column=2).value == 300

        # Row 5-6: devices
        assert ws.cell(row=5, column=1).value == "ROOM A"
        assert ws.cell(row=5, column=2).value == 100
        assert ws.cell(row=6, column=1).value == "ROOM B"
        assert ws.cell(row=6, column=2).value == 200

        # Row 7: grand total
        assert ws.cell(row=7, column=1).value == "Totale complessivo"
        assert ws.cell(row=7, column=2).value == 300


def test_raw_sheet_header():
    """Raw sheet should have correct file header."""
    report, sorted_hca, pivot_groups = _make_report()

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "test.xlsx"
        write_xlsx(report, sorted_hca, pivot_groups, out)
        wb = openpyxl.load_workbook(out)
        ws = wb[wb.sheetnames[1]]

        assert ws.cell(row=1, column=1).value == "Nome file"
        assert ws.cell(row=2, column=1).value == "RAW_report_9999_2026-01-01_09-00"
        assert ws.cell(row=2, column=4).value == "TEST BUILDING"


def test_raw_sheet_sum_formulas():
    """Raw sheet should have SUM formulas at the bottom."""
    report, sorted_hca, pivot_groups = _make_report()

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "test.xlsx"
        write_xlsx(report, sorted_hca, pivot_groups, out)
        wb = openpyxl.load_workbook(out)
        ws = wb[wb.sheetnames[1]]

        # Find the last row with data
        last_row = ws.max_row
        # Column M (13) should have a SUM formula
        val = ws.cell(row=last_row, column=13).value
        assert val is not None and "SUM" in str(val)


# -- Style tests ---------------------------------------------------------------


def _write_test_workbook():
    """Helper: write workbook and return opened sheets."""
    report, sorted_hca, pivot_groups = _make_report()
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "test.xlsx"
        write_xlsx(report, sorted_hca, pivot_groups, out)
        wb = openpyxl.load_workbook(out)
        yield wb
        wb.close()


def test_pivot_header_has_dark_fill():
    """PIVOT header row should have dark blue background and bold white text."""
    for wb in _write_test_workbook():
        ws = wb["PIVOT "]
        cell = ws.cell(row=3, column=1)
        assert cell.font.bold is True
        assert cell.fill.start_color.rgb == "002F5496"
        assert cell.font.color.rgb == "00FFFFFF"


def test_pivot_apartment_total_has_accent_fill():
    """PIVOT apartment total rows should be light blue and bold."""
    for wb in _write_test_workbook():
        ws = wb["PIVOT "]
        cell = ws.cell(row=4, column=1)
        assert cell.font.bold is True
        assert cell.fill.start_color.rgb == "00D6E4F0"


def test_pivot_grand_total_has_dark_fill():
    """PIVOT grand total row should have dark blue background and bold white text."""
    for wb in _write_test_workbook():
        ws = wb["PIVOT "]
        last_row = ws.max_row
        cell = ws.cell(row=last_row, column=1)
        assert cell.value == "Totale complessivo"
        assert cell.font.bold is True
        assert cell.fill.start_color.rgb == "002F5496"


def test_pivot_freeze_panes():
    """PIVOT sheet should freeze panes at A4."""
    for wb in _write_test_workbook():
        ws = wb["PIVOT "]
        assert str(ws.freeze_panes) == "A4"


def test_raw_header_has_dark_fill():
    """RAW file header row should have dark blue background."""
    for wb in _write_test_workbook():
        ws = wb[wb.sheetnames[1]]
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.fill.start_color.rgb == "002F5496"


def test_raw_hca_header_has_dark_fill():
    """RAW HCA column header row should have dark blue background."""
    for wb in _write_test_workbook():
        ws = wb[wb.sheetnames[1]]
        # With 1 concentrator: rows 4-6 (header/data/blank), row 9 subheader,
        # row 10 HCA header. But our data has row structure:
        # 1=file header, 2=file data, 3=blank, 4=conc header, 5=conc data,
        # 6=blank, 7=subheader (Attuale/Es.Prec), 8=HCA header, 9+=HCA data
        # Find the HCA header row by looking for "count" with dark fill
        for row in range(4, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value == "count" and cell.fill.start_color.rgb == "002F5496":
                # Found the HCA header row with dark fill
                assert cell.font.bold is True
                return
        raise AssertionError("No HCA header row with dark fill found")


def test_raw_freeze_panes():
    """RAW sheet should have freeze panes set."""
    for wb in _write_test_workbook():
        ws = wb[wb.sheetnames[1]]
        assert ws.freeze_panes is not None


def test_raw_autofilter_has_sort_state():
    """RAW sheet autofilter should sort by device_detail (col F) descending, then device_description (col E) ascending."""
    for wb in _write_test_workbook():
        ws = wb[wb.sheetnames[1]]
        sort_state = ws.auto_filter.sortState
        assert sort_state is not None, "autofilter has no sortState"
        conditions = sort_state.sortCondition
        refs = [sc.ref for sc in conditions]
        # First sort key: column F (device_detail) — descending
        assert any(r.startswith("F") for r in refs), f"No sort on column F, got: {refs}"
        f_idx = next(i for i, r in enumerate(refs) if r.startswith("F"))
        assert conditions[f_idx].descending is True, (
            "device_detail (F) must sort descending"
        )
        # Second sort key: column E (device_description) — ascending
        assert any(r.startswith("E") for r in refs), f"No sort on column E, got: {refs}"
        e_idx = next(i for i, r in enumerate(refs) if r.startswith("E"))
        assert f_idx < e_idx, (
            "device_detail (F) must be primary sort, device_description (E) secondary"
        )


def test_raw_columns_auto_fitted():
    """RAW sheet columns should have non-default widths (auto-fitted)."""
    for wb in _write_test_workbook():
        ws = wb[wb.sheetnames[1]]
        # Column A should be wider than default (8) since it has content
        width_a = ws.column_dimensions["A"].width
        assert width_a is not None and width_a > 8
